#!/usr/bin/env python
# coding: utf-8

import pandas as pd
from openpyxl import load_workbook
import datetime
import os


def write_dayuse_to_excel(shtname, dfname):
    alist = dfname
    ws = wb2[shtname]
    # cell_range = ws['E21':'E51']
    i = 0
    while i < len(alist):
        ws.cell(row=i + 21, column=5, value=alist[i])
        i += 1


def write_day_max_demand(shtname, dfname):
    alist = dfname
    ws = wb2[shtname]
    # cell_range = ws['C21':'C51']
    i = 0
    while i < len(alist):
        ws.cell(row=i + 21, column=3, value=alist[i])
        i += 1


def write_date(shtname):
    ws = wb2[shtname]
    ws['D6'] = datetime.datetime(now.year, now.month-1, 1).strftime('%b %Y')


def write_start_value(shtname, startvalue):
    ws = wb2[shtname]
    ws['E9'] = startvalue
    # ws['E8'] = endvalue


def write_end_value(shtname, endvalue):
    ws = wb2[shtname]
    ws['E8'] = endvalue


def write_start_date(shtname):
    ws = wb2[shtname]
    ws['B8'] = datetime.datetime(now.year, now.month-1, 1).strftime('%d.%m.%y')


def write_end_date(shtname):
    ws = wb2[shtname]
    ws['B9'] = datetime.datetime(now.year, now.month, 1).strftime('%d.%m.%y')


pd.options.mode.chained_assignment = None

current_date = datetime.datetime.now().strftime('%Y%m')  # 获取当前年月格式2023xx
now = now = datetime.datetime.now()
last_month = datetime.datetime(now.year, now.month-1, 1).strftime('%Y%m')
# monthfirstday = datetime.datetime.now().strftime('%Y%m%d')
# monthfirstday
today = datetime.datetime.today()
# monthfirstday = datetime.datetime(today.year, today.month, 1, 0, 0, 0).strftime('%Y%m%d %H:%M:%S')

path1 = r"C:\Users\hank\PythonCode\Monthlybill"
files = os.listdir(path1)
pattern = 'CHW RTU_Monthly Report_'+current_date
for f in files:
    if pattern in f:
        path = f
# 模糊查找，文件夹下含有'CHW RTU_Monthly Report_'+current_date的文件
filepath = path1+"\\"+path

# 需要读取上一月的数据
pattern_lastMonth = 'CHW RTU_Monthly Report_'+last_month
for f in files:
    if pattern_lastMonth in f:
        path_last_month = f
filepath_lastmonth = path1+"\\"+path_last_month

df = pd.read_excel(filepath,
                   sheet_name='Report',
                   header=1,
                   parse_dates=['datetime']
                   )
df = df.dropna()  # 删除NaN
df_lastmonth = pd.read_excel(filepath_lastmonth,
                             sheet_name='Report',
                             header=1,
                             parse_dates=['datetime']
                             )
df_lastmonth = df_lastmonth.dropna()  # 删除NaN


df = pd.concat([df_lastmonth.tail(1), df[1:]])  # 用上月最后一行替换当月第一行
df.reset_index(drop=True, inplace=True)


# 获取每日最大负荷QI
df_max_daily_QI = df.groupby(pd.Grouper(key='datetime', axis=0,
                                        freq='D')).max()  # 用于每天的最大负荷需求
df_max_daily_QI.drop(df_max_daily_QI.tail(1).index, inplace=True)  # 删除最后一行

df.drop(df.columns[9:105], axis=1, inplace=True)  # 删除后面的列
# df.drop(df.index[0], inplace=True)  # 删除第一行
daymin = df.groupby(pd.Grouper(key='datetime', axis=0,
                               freq='D')).min()
daymin_s = daymin[:-1]
daymin_e = daymin[1:]


data_0 = df[df["datetime"].dt.hour == 0]  # .dt获取每天某小时的数据
data_0_s = data_0[:-1]
data_0_e = data_0[1:]
data_0_e.reset_index(inplace=True, drop=True)
data_0_e.index = data_0_e.index + 1
data_0_s.reset_index(
    inplace=True, drop=True
)  # 为了方便用pandas的减法，需要将index重置，这里用到reset_index
# data_0_e=data_0_e.drop(columns=[Index])
# data_0_s=data_0_s.drop(columns=[Index])
data_0_s.index = data_0_s.index + 1
dayuse = (
    data_0_e
    - data_0_s
)
# dayuse.drop(dayuse['datetime'],axis=1,inplace=True)
# dayuse.index = daymin_s.index


dayuse.reset_index(drop=True, inplace=True)
dayuse.drop(columns=['datetime'], axis=1, inplace=True)
dayuse.set_index(daymin_s.index, inplace=True, verify_integrity=True)


wb2 = load_workbook("Monthlybill\RCCQ CHW Metering Report.xlsx")
# sheet_name = wb2.sheetnames
# print(sheet_name)


write_date('Summary')
write_end_date('PodiumE')
write_start_date('PodiumE')
# worksheet = wb2['PodiumE']
# A1 = worksheet['E9']
# print(A1.value)
write_start_value('PodiumE', df_lastmonth.D1_QQ_1150_BTU.max())
write_start_value('PodiumW', df_lastmonth.D2_QQ_1150_BTU.max())
write_start_value('HotelHZ', df_lastmonth.D3_QQ_1150_BTU.max())
write_start_value('HotelLZ', df_lastmonth.D4_QQ_1150_BTU.max())
write_start_value('OfficeHZ', df_lastmonth.D5_QQ_1150_BTU.max())
write_start_value('OfficeLZ', df_lastmonth.D6_QQ_1150_BTU.max())
write_start_value('SvcAptHZ', df_lastmonth.D7_QQ_1150_BTU.max())
write_start_value('SvcAptLZ', df_lastmonth.D8_QQ_1150_BTU.max())

write_end_value('PodiumE', df.D1_QQ_1150_BTU.max())
write_end_value('PodiumW',  df.D2_QQ_1150_BTU.max())
write_end_value('HotelHZ',  df.D3_QQ_1150_BTU.max())
write_end_value('HotelLZ', df.D4_QQ_1150_BTU.max())
write_end_value('OfficeHZ', df.D5_QQ_1150_BTU.max())
write_end_value('OfficeLZ',  df.D6_QQ_1150_BTU.max())
write_end_value('SvcAptHZ',  df.D7_QQ_1150_BTU.max())
write_end_value('SvcAptLZ',  df.D8_QQ_1150_BTU.max())


write_dayuse_to_excel('PodiumE', dayuse['D1_QQ_1150_BTU'])
write_dayuse_to_excel('PodiumW', dayuse['D2_QQ_1150_BTU'])
write_dayuse_to_excel('HotelHZ', dayuse['D3_QQ_1150_BTU'])
write_dayuse_to_excel('HotelLZ', dayuse['D4_QQ_1150_BTU'])
write_dayuse_to_excel('OfficeHZ', dayuse['D5_QQ_1150_BTU'])
write_dayuse_to_excel('OfficeLZ', dayuse['D6_QQ_1150_BTU'])
write_dayuse_to_excel('SvcAptHZ', dayuse['D7_QQ_1150_BTU'])
write_dayuse_to_excel('SvcAptLZ', dayuse['D8_QQ_1150_BTU'])

write_day_max_demand('PodiumE', df_max_daily_QI['D1_QI_1150_BTU'])
write_day_max_demand('PodiumW', df_max_daily_QI['D2_QI_1150_BTU'])
write_day_max_demand('HotelHZ', df_max_daily_QI['D3_QI_1150_BTU'])
write_day_max_demand('HotelLZ', df_max_daily_QI['D4_QI_1150_BTU'])
write_day_max_demand('OfficeHZ', df_max_daily_QI['D5_QI_1150_BTU'])
write_day_max_demand('OfficeLZ', df_max_daily_QI['D6_QI_1150_BTU'])
write_day_max_demand('SvcAptHZ', df_max_daily_QI['D7_QI_1150_1_BTU'])
write_day_max_demand('SvcAptLZ', df_max_daily_QI['D8_QI_1150_2_BTU'])

wb2.save("RCCQ CHW Metering Report.xlsx")
